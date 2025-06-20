import sys
import os
import requests
import getopt
import urllib.parse
import openpyxl
from veracode_api_py.applications import Applications
from veracode_api_py.identity import BusinessUnits, Teams
from veracode_api_py.policy import Policies

from veracode_api_signing.credentials import get_credentials

class NoExactMatchFoundException(Exception):
    message=""
    def __init__(self, message_to_set):
        self.message = message_to_set

    def get_message(self):
        return self.message

NULL = "NULL"

json_headers = {
    "User-Agent": "Bulk application creation - python script",
    "Content-Type": "application/json"
}

failed_attempts = 0
max_attempts_per_request = 10
sleep_time = 10

last_column = 0
non_custom_field_headers={"Application Name",
                          "New Application Name",
                            "Business Criticality",
                            "Policy",
                            "Description",
                            "Tags",
                            "Business Unit",
                            "Business Owner",
                            "Owner Email",
                            "Teams"}

def print_help():
    """Prints command line options and exits"""
    print("""bulk-update-applications.py -f <excel_file_with_application_definitions> [-r <header_row>] [-d]"
        Reads all lines in <excel_file_with_application_definitions>, for each line, it will update the profile
        <header_row> defines which row contains your table headers, which will be read to determine where each field goes (default 2).
""")
    sys.exit()

def get_field_value(excel_headers, excel_sheet, row, field_header):
    field_to_get = field_header.strip()
    if field_to_get in excel_headers:
        field_value = excel_sheet.cell(row = row, column = excel_headers[field_to_get]).value
        if field_value:
            return field_value
    return ""

def get_business_owners(excel_headers, excel_sheet, row):
    name=get_field_value(excel_headers, excel_sheet, row, "Business Owner")
    email=get_field_value(excel_headers, excel_sheet, row, "Owner Email")
    if not name or not email:
        return None
    elif name == NULL or email == NULL:
        return ""
    else:
        return { "email": email, "name": name}

def get_business_unit_guid(business_unit_name: str):
    matches = BusinessUnits().get_all()
    if not matches or len(matches) == 0:
        return None
    for match in matches:
        if match["bu_name"].strip() == business_unit_name.strip():
            return match["bu_id"]
    return None

def get_business_unit(excel_headers, excel_sheet, row):
    business_unit_name = get_field_value(excel_headers, excel_sheet, row, "Business Unit")
    if not business_unit_name:
        return None
    elif business_unit_name == NULL:
        return ""
    else:
        bu_guid = get_business_unit_guid(business_unit_name)
        if not bu_guid:
            raise NoExactMatchFoundException(f"Unable to find a Business Unit named {business_unit_name}")
        return bu_guid

def get_policy_guid(policy_name: str):
    matches = Policies().get_all()
    if not matches or len(matches) == 0:
        return None
    for match in matches:
        if match["name"].strip() == policy_name.strip():
            return match["guid"]
    return None

def get_policy(excel_headers, excel_sheet, row):
    policy_name = get_field_value(excel_headers, excel_sheet, row, "Policy")
    if not policy_name:
        return None
    else:
        policy_guid = get_policy_guid(policy_name)
        if not policy_guid:
            raise NoExactMatchFoundException(f"Unable to find a Policy named {policy_name}")
        return policy_guid

def inner_get_team_guid(team_name: str):
    matches = Teams().get_all()
    if not matches or len(matches) == 0:
        return None
    for match in matches:
        if match["team_name"].strip() == team_name.strip():
            return match["team_id"]
    return None

def get_team_guid(team_name):
    team_guid = inner_get_team_guid(team_name)
    if not team_guid:
        raise NoExactMatchFoundException(f"Unable to find a Team named {team_name}")
    return team_guid

def get_teams(excel_headers, excel_sheet, row):
    all_teams_base = get_field_value(excel_headers, excel_sheet, row, "Teams")
    if not all_teams_base:
        return []
    all_teams = all_teams_base.split(",")
    inner_team_list = []
    for team_name in all_teams:
        inner_team_list.append(get_team_guid(team_name.strip()))
    return inner_team_list

def parse_custom_field_list(original_list, new_list):
    new_list = list(map(lambda custom_field: parse_custom_field(custom_field), new_list))
    all_names = list(map(lambda custom_field: custom_field["name"], new_list))
    for field in original_list:
        if not field["name"] in all_names:
            new_list.append(field)
    return new_list

def parse_new_custom_fields(excel_headers, excel_sheet, row):
    global non_custom_field_headers
    inner_custom_fields_list = []
    for field in excel_headers:
        if not field in non_custom_field_headers:
            value = excel_sheet.cell(row = row, column=excel_headers[field]).value
            if value:
                inner_custom_fields_list.append({
                    "name": field,
                    "value": value
                })
    if inner_custom_fields_list:
        return inner_custom_fields_list
    else:
        return []
    
def parse_custom_field(custom_field):
    return {
        "name": custom_field["name"],
        "value": custom_field["value"]
    }

def get_description(excel_headers, excel_sheet, row):
    description = get_field_value(excel_headers, excel_sheet, row, "Description")
    if description == NULL:
        return ""
    if not description:
        return None
    return description        

def get_tags(excel_headers, excel_sheet, row):
    tags = get_field_value(excel_headers, excel_sheet, row, "Tags")
    if tags == NULL:
        return ""
    if not tags:
        return None
    return tags

def get_business_criticality(excel_headers, excel_sheet, row):
    business_criticality = get_field_value(excel_headers, excel_sheet, row, "Business Criticality").replace(" ", "_").upper()
    if not business_criticality:
        return None
    return business_criticality

def get_current_business_owner(application):
    if not "business_owners" in application["profile"]:
        return None
    return application["profile"]["business_owners"]

def get_current_business_unit(application):
    if not "business_unit" in application["profile"]:
        return None
    return application["profile"]["business_unit"]["guid"]

def get_current_description(application):
    if not "description" in application["profile"]:
        return None
    return application["profile"]["description"]

def get_current_policy(application):
    if not "policies" in application["profile"] or not application["profile"]["policies"]:
        return None
    return application["profile"]["policies"][0]["guid"]

def get_current_tags(application):
    if not "tags" in application["profile"]:
        return None
    return application["profile"]["tags"]

def get_current_teams(application):
    if not "teams" in application["profile"]:
        return None
    
    inner_team_list = []
    for team in application["profile"]["teams"]:
        inner_team_list.append(team["guid"])
    return inner_team_list

def try_update_application(application, new_name, excel_headers, excel_sheet, row):
    business_criticality_new = get_business_criticality(excel_headers, excel_sheet, row)
    if not business_criticality_new:
        business_criticality_new = application["profile"]["business_criticality"]
    business_owner_new = get_business_owners(excel_headers, excel_sheet, row)
    business_unit_new = get_business_unit(excel_headers, excel_sheet, row)
    description_new = get_description(excel_headers, excel_sheet, row)
    policy_new= get_policy(excel_headers, excel_sheet, row)
    tags_new = get_tags(excel_headers, excel_sheet, row)
    teams_new = get_teams(excel_headers, excel_sheet, row)

    business_owner_new = get_current_business_owner(application) if business_owner_new == None else business_owner_new
    business_unit_new = get_current_business_unit(application) if business_unit_new == None else business_unit_new
    description_new = get_current_description(application) if description_new == None else description_new
    policy_new = get_current_policy(application) if policy_new == None else policy_new
    tags_new = get_current_tags(application) if tags_new == None else tags_new
    teams_new = get_current_teams(application) if teams_new == None else teams_new



    new_custom_fields = parse_new_custom_fields(excel_headers, excel_sheet, row)
    if new_custom_fields:
        current_custom_fields = list(map(lambda custom_field: custom_field, application["profile"]["custom_fields"])) if application["profile"]["custom_fields"] else []
        custom_fields_new=parse_custom_field_list(current_custom_fields, new_custom_fields)
    else:
        custom_fields_new = None
    
    git_repo_url = application["profile"]["git_repo_url"] if "git_repo_url" in application["profile"] else None
    custom_kms_alias= application["profile"]["custom_kms_alias"] if "custom_kms_alias" in application["profile"] else None

    Applications().update(guid=application["guid"], app_name=new_name, business_criticality=business_criticality_new, description=description_new, business_unit=business_unit_new, 
                          teams=teams_new, policy_guid=policy_new, custom_fields=custom_fields_new, bus_owner_name=business_owner_new['name'] if business_owner_new else None, bus_owner_email=business_owner_new['email'] if business_owner_new else None, 
                          tags=tags_new, custom_kms_alias=custom_kms_alias, git_repo_url=git_repo_url)
    return "success"

def get_application(application_name: str):
    matches = Applications().get_by_name(application_name)
    if not matches or len(matches) == 0:
        return None
    for match in matches:
        if match["profile"]["name"] == application_name.strip():
            return match
    return None

def update_application(excel_headers, excel_sheet, row):
    application_name = get_field_value(excel_headers, excel_sheet, row, "Application Name")
    new_name = get_field_value(excel_headers, excel_sheet, row, "New Application Name")
    if not new_name:
        new_name = application_name

    application = get_application(application_name)
    if not application:
        error_message = f"Application not found: {application_name}"
        print (error_message)
        return error_message
    
    return try_update_application(application, new_name, excel_headers, excel_sheet, row)

def setup_excel_headers(excel_sheet, header_row, verbose):
    excel_headers = {}
    global last_column
    for column in range(1, excel_sheet.max_column+1):
        cell = excel_sheet.cell(row = header_row, column = column)
        if not cell or cell is None or not cell.value or cell.value.strip() == "":
            break
        to_add = cell.value
        if to_add:
            to_add = str(to_add).strip()
        if verbose:
            print(f"Adding column {column} for value {to_add}")
        excel_headers[to_add] = column
        last_column += 1
    return excel_headers

def update_all_applications(file_name, header_row, verbose):
    global failed_attempts
    excel_file = openpyxl.load_workbook(file_name)
    excel_sheet = excel_file.active
    try:
        excel_headers = setup_excel_headers(excel_sheet, header_row, verbose)
        print("Finished reading excel headers")
        if verbose:
            print("Values found are:")
            print(excel_headers)

        max_column=len(excel_headers)
        for row in range(header_row+1, excel_sheet.max_row+1):      
            failed_attempts = 0
            if verbose:
                for field in excel_headers:
                    print(f"Found column with values:")
                    print(f"{field} -> {excel_sheet.cell(row = row, column=excel_headers[field]).value}")
            status=excel_sheet.cell(row = row, column = max_column+1).value
            if (status == 'success'):
                print("Skipping row as it was already done")
            else:
                try:
                    print(f"Importing row {row-header_row}/{excel_sheet.max_row-header_row}:")
                    status = update_application(excel_headers, excel_sheet, row)
                    print(f"Finished importing row {row-header_row}/{excel_sheet.max_row-header_row}")
                    print("---------------------------------------------------------------------------")
                except NoExactMatchFoundException as nemfe:
                    status= nemfe.get_message()
                except Exception as e:
                    status = repr(e)
                excel_sheet.cell(row = row, column = max_column+1).value=status
    finally:
        excel_file.save(filename=file_name)

def main(argv):
    """Allows for bulk updating application profiles"""
    global failed_attempts
    global last_column
    excel_file = None

    old_lower_veracode_api_key_id = os.environ.get('veracode_api_key_id', "")
    old_lower_veracode_api_key_secret = os.environ.get('veracode_api_key_secret', "")
    old_upper_veracode_api_key_id = os.environ.get('VERACODE_API_KEY_ID', "")
    old_upper_veracode_api_key_secret = os.environ.get('VERACODE_API_KEY_SECRET', "")

    try:
        verbose = False
        file_name = ''
        header_row = -1

        opts, args = getopt.getopt(argv, "hdf:r:", ["file_name=", "header_row="])
        for opt, arg in opts:
            if opt == '-h':
                print_help()
            if opt == '-d':
                verbose = True
            if opt in ('-f', '--file_name'):
                file_name=arg
            if opt in ('-r', '--header_row'):
                header_row=int(arg)

        if not file_name:
            print_help()
        if header_row < 0:
            print("INFO: No header row set, using default of 2")
            header_row = 2
        
        update_all_applications(file_name, header_row, verbose)
    except requests.RequestException as e:
        print("An error occurred!")
        print(e)
        sys.exit(1)
    finally:
        os.environ['veracode_api_key_id'] = old_lower_veracode_api_key_id
        os.environ['VERACODE_API_KEY_ID'] = old_upper_veracode_api_key_id
        os.environ['veracode_api_key_secret'] = old_lower_veracode_api_key_secret
        os.environ['VERACODE_API_KEY_SECRET'] = old_upper_veracode_api_key_secret
        if excel_file:
            excel_file.save(filename=file_name)


if __name__ == "__main__":
    main(sys.argv[1:])
